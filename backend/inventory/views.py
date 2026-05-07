from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import csv
import io

from .models import (
    Category, Unit, MedicationStock, StockBatch, StockAdjustment,
    InventoryCount, InventoryCountLine,
    StockTransfer, StockTransferLine,
    ControlledSubstanceLog,
)
from .serializers import (
    CategorySerializer, UnitSerializer,
    MedicationStockSerializer, StockBatchSerializer, StockAdjustmentSerializer,
    InventoryCountSerializer, InventoryCountLineSerializer,
    StockTransferSerializer,
    ControlledSubstanceLogSerializer,
)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['name', 'created_at']


class UnitViewSet(viewsets.ModelViewSet):
    queryset = Unit.objects.all()
    serializer_class = UnitSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'abbreviation']
    ordering_fields = ['name', 'created_at']


class MedicationStockViewSet(viewsets.ModelViewSet):
    queryset = MedicationStock.objects.select_related('category', 'unit').prefetch_related('batches').all()
    serializer_class = MedicationStockSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'category', 'unit']
    search_fields = ['medication_name', 'abbreviation', 'barcode']
    ordering_fields = ['medication_name', 'selling_price', 'created_at']

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        stocks = self.get_queryset().filter(is_active=True)
        low = [s for s in stocks if s.is_low_stock]
        serializer = self.get_serializer(low, many=True)
        return Response(serializer.data)

    # ── Bulk operations ───────────────────────────────────────────────────
    _BULK_EDITABLE_FIELDS = {
        'medication_name', 'category', 'unit',
        'selling_price', 'cost_price', 'discount_percent',
        'reorder_level', 'reorder_quantity',
        'location_in_store', 'barcode',
        'prescription_required', 'is_active',
    }

    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update(self, request):
        """Update many stock items at once.

        Payload: { "items": [{"id": 1, "selling_price": 12.5, ...}, ...] }
        Only fields in _BULK_EDITABLE_FIELDS are accepted; unknown fields ignored.
        """
        items = request.data.get('items') or []
        if not isinstance(items, list) or not items:
            return Response({'detail': 'items must be a non-empty list.'}, status=400)

        ids = [i.get('id') for i in items if i.get('id')]
        qs = MedicationStock.objects.filter(id__in=ids)
        by_id = {s.id: s for s in qs}

        updated = []
        errors = []
        for payload in items:
            sid = payload.get('id')
            obj = by_id.get(sid)
            if not obj:
                errors.append({'id': sid, 'error': 'not found'})
                continue
            cleaned = {k: v for k, v in payload.items() if k in self._BULK_EDITABLE_FIELDS}

            # Optional inline quantity edit (creates a count-correction adjustment).
            qty_error = None
            if 'set_quantity' in payload and payload['set_quantity'] is not None:
                try:
                    new_qty = int(payload['set_quantity'])
                    if new_qty < 0:
                        raise ValueError('quantity must be ≥ 0')
                    self._apply_quantity_set(obj, new_qty, request.user)
                except (TypeError, ValueError) as exc:
                    qty_error = str(exc) or 'invalid quantity'

            if qty_error:
                errors.append({'id': sid, 'error': {'set_quantity': [qty_error]}})

            if cleaned:
                serializer = self.get_serializer(obj, data=cleaned, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    updated.append(serializer.data)
                else:
                    errors.append({'id': sid, 'error': serializer.errors})
            elif 'set_quantity' in payload and not qty_error:
                updated.append(self.get_serializer(obj).data)

        return Response({'updated': len(updated), 'items': updated, 'errors': errors})

    def _apply_quantity_set(self, stock, new_qty, user):
        """Reconcile total quantity to ``new_qty`` via a StockAdjustment.

        For increases without any existing batch, a synthetic batch is created
        (1-year expiry). For decreases, batches are drained FEFO.
        """
        from datetime import date, timedelta as _td
        current = stock.total_quantity
        delta = new_qty - current
        if delta == 0:
            return
        if delta > 0:
            batch = stock.batches.filter(quantity_remaining__gt=0).order_by('-expiry_date').first()
            if batch is None:
                batch = stock.batches.order_by('-expiry_date').first()
            if batch is None:
                batch = StockBatch.objects.create(
                    stock=stock,
                    batch_number=f'ADJ-{stock.pk}-{date.today():%Y%m%d}',
                    quantity_received=delta,
                    quantity_remaining=0,
                    cost_price_per_unit=stock.cost_price or 0,
                    expiry_date=date.today() + _td(days=365),
                )
            StockAdjustment.objects.create(
                stock=stock, batch=batch, quantity_change=delta,
                reason=StockAdjustment.Reason.COUNT_CORRECTION,
                notes='Bulk edit count correction',
                adjusted_by=user if user.is_authenticated else None,
            )
            batch.quantity_remaining = batch.quantity_remaining + delta
            batch.save(update_fields=['quantity_remaining'])
        else:
            remaining = -delta
            for batch in stock.batches.filter(quantity_remaining__gt=0).order_by('expiry_date'):
                if remaining <= 0:
                    break
                take = min(batch.quantity_remaining, remaining)
                StockAdjustment.objects.create(
                    stock=stock, batch=batch, quantity_change=-take,
                    reason=StockAdjustment.Reason.COUNT_CORRECTION,
                    notes='Bulk edit count correction',
                    adjusted_by=user if user.is_authenticated else None,
                )
                batch.quantity_remaining -= take
                batch.save(update_fields=['quantity_remaining'])
                remaining -= take

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """Delete many stock items at once. Payload: { \"ids\": [1,2,3] }"""
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'detail': 'ids must be a non-empty list.'}, status=400)
        qs = MedicationStock.objects.filter(id__in=ids)
        count = qs.count()
        qs.delete()
        return Response({'deleted': count, 'ids': ids})

    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        days = int(request.query_params.get('days', 30))
        cutoff = timezone.now().date() + timedelta(days=days)
        batches = StockBatch.objects.filter(
            expiry_date__lte=cutoff,
            quantity_remaining__gt=0,
        ).select_related('stock').order_by('expiry_date')
        serializer = StockBatchSerializer(batches, many=True)
        return Response(serializer.data)

    # ── Export ────────────────────────────────────────────────────────────────
    _EXPORT_HEADERS = [
        'ID', 'Medication Name', 'Category', 'Unit',
        'Selling Price (KSh)', 'Cost Price (KSh)',
        'Total Quantity', 'Reorder Level', 'Reorder Quantity',
        'Location', 'Barcode', 'Prescription Required', 'Active',
        'Created At',
    ]

    def _stock_row(self, s):
        return [
            s.id, s.medication_name,
            s.category.name if s.category else '',
            s.unit.name if s.unit else '',
            float(s.selling_price), float(s.cost_price),
            s.total_quantity, s.reorder_level, s.reorder_quantity,
            s.location_in_store or '', s.barcode or '',
            s.prescription_required, 'Yes' if s.is_active else 'No',
            s.created_at.strftime('%Y-%m-%d'),
        ]

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        qs = self.filter_queryset(self.get_queryset())

        if fmt == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                return Response({'detail': 'openpyxl not installed.'}, status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Inventory'

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(fill_type='solid', fgColor='1565C0')
            for col_idx, header in enumerate(self._EXPORT_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row_idx, s in enumerate(qs, 2):
                for col_idx, val in enumerate(self._stock_row(s), 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f'inventory_{timezone.now().strftime("%Y%m%d")}.xlsx'
            return HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{fname}"'},
            )

        elif fmt == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
                from reportlab.lib.units import cm
            except ImportError:
                return Response({'detail': 'reportlab not installed.'}, status=500)

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
            styles = getSampleStyleSheet()

            elements = [
                Paragraph('Inventory Report', styles['Title']),
                Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']),
                Spacer(1, 0.4*cm),
            ]

            pdf_headers = ['Name', 'Category', 'Unit', 'Sell Price', 'Cost', 'Qty', 'Reorder', 'Location', 'Barcode', 'Rx', 'Active']
            data = [pdf_headers]
            for s in qs:
                data.append([
                    s.medication_name,
                    s.category.name if s.category else '',
                    s.unit.abbreviation if s.unit else '',
                    f'KSh {float(s.selling_price):.2f}',
                    f'KSh {float(s.cost_price):.2f}',
                    str(s.total_quantity),
                    str(s.reorder_level),
                    s.location_in_store or '',
                    s.barcode or '',
                    s.prescription_required,
                    'Yes' if s.is_active else 'No',
                ])

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(table)
            doc.build(elements)
            buf.seek(0)
            fname = f'inventory_{timezone.now().strftime("%Y%m%d")}.pdf'
            return HttpResponse(
                buf.read(),
                content_type='application/pdf',
                headers={'Content-Disposition': f'attachment; filename="{fname}"'},
            )

        else:
            # Default: CSV
            fname = f'inventory_{timezone.now().strftime("%Y%m%d")}.csv'
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{fname}"'
            writer = csv.writer(response)
            writer.writerow(self._EXPORT_HEADERS)
            for s in qs:
                writer.writerow(self._stock_row(s))
            return response

    # ── Import ────────────────────────────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='import')
    def import_stocks(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = uploaded.name.lower()
        created = updated = errors = 0
        error_details = []

        def _process_row(row_num, name, category_name, selling_price_str, cost_price_str, reorder_level_str, reorder_qty_str, location, barcode, prescription):
            nonlocal created, updated, errors
            name = (name or '').strip()
            if not name:
                errors += 1
                error_details.append(f'Row {row_num}: medication name is required.')
                return
            try:
                selling_price = float(selling_price_str or 0)
                cost_price = float(cost_price_str or 0)
                reorder_level = int(float(reorder_level_str or 10))
                reorder_qty = int(float(reorder_qty_str or 20))
            except (ValueError, TypeError):
                errors += 1
                error_details.append(f'Row {row_num}: invalid numeric value.')
                return

            category = None
            if category_name and category_name.strip():
                category, _ = Category.objects.get_or_create(name=category_name.strip())

            valid_rx = {'none', 'recommended', 'required'}
            rx = (prescription or 'none').strip().lower()
            if rx not in valid_rx:
                rx = 'none'

            stock, was_created = MedicationStock.objects.update_or_create(
                medication_name=name,
                defaults=dict(
                    category=category,
                    selling_price=selling_price,
                    cost_price=cost_price,
                    reorder_level=reorder_level,
                    reorder_quantity=reorder_qty,
                    location_in_store=(location or '').strip() or None,
                    barcode=(barcode or '').strip() or None,
                    prescription_required=rx,
                ),
            )
            if was_created:
                created += 1
            else:
                updated += 1

        if filename.endswith('.csv'):
            try:
                text = uploaded.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                for i, row in enumerate(reader, 2):
                    _process_row(
                        i,
                        row.get('Medication Name') or row.get('medication_name', ''),
                        row.get('Category') or row.get('category', ''),
                        row.get('Selling Price (KSh)') or row.get('selling_price', ''),
                        row.get('Cost Price (KSh)') or row.get('cost_price', ''),
                        row.get('Reorder Level') or row.get('reorder_level', ''),
                        row.get('Reorder Quantity') or row.get('reorder_quantity', ''),
                        row.get('Location') or row.get('location', ''),
                        row.get('Barcode') or row.get('barcode', ''),
                        row.get('Prescription Required') or row.get('prescription_required', ''),
                    )
            except Exception as e:
                return Response({'detail': f'CSV parse error: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        elif filename.endswith(('.xlsx', '.xls')):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return Response({'detail': 'openpyxl not installed.'}, status=500)
            try:
                wb = load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return Response({'detail': 'Empty spreadsheet.'}, status=status.HTTP_400_BAD_REQUEST)
                headers = [str(h).strip() if h else '' for h in rows[0]]
                col = {h: i for i, h in enumerate(headers)}

                def _get(row, key, fallback=''):
                    idx = col.get(key)
                    return str(row[idx] or '') if idx is not None and idx < len(row) else fallback

                for i, row in enumerate(rows[1:], 2):
                    _process_row(
                        i,
                        _get(row, 'Medication Name') or _get(row, 'medication_name'),
                        _get(row, 'Category') or _get(row, 'category'),
                        _get(row, 'Selling Price (KSh)') or _get(row, 'selling_price'),
                        _get(row, 'Cost Price (KSh)') or _get(row, 'cost_price'),
                        _get(row, 'Reorder Level') or _get(row, 'reorder_level'),
                        _get(row, 'Reorder Quantity') or _get(row, 'reorder_quantity'),
                        _get(row, 'Location') or _get(row, 'location'),
                        _get(row, 'Barcode') or _get(row, 'barcode'),
                        _get(row, 'Prescription Required') or _get(row, 'prescription_required'),
                    )
            except Exception as e:
                return Response({'detail': f'Excel parse error: {e}'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'detail': 'Unsupported file type. Use .csv or .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'error_details': error_details[:20],
        })


class StockBatchViewSet(viewsets.ModelViewSet):
    queryset = StockBatch.objects.select_related('stock', 'supplier').all()
    serializer_class = StockBatchSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['stock', 'expiry_date']
    search_fields = ['batch_number', 'stock__medication_name']
    ordering_fields = ['expiry_date', 'received_date']


class StockAdjustmentViewSet(viewsets.ModelViewSet):
    queryset = StockAdjustment.objects.select_related('stock', 'batch', 'adjusted_by').all()
    serializer_class = StockAdjustmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['stock', 'reason']
    search_fields = ['stock__medication_name', 'notes']
    ordering_fields = ['created_at']


class InventoryAnalyticsView(APIView):
    """Aggregated inventory analytics for the pharmacy dashboard."""

    def get(self, request):
        stocks = MedicationStock.objects.prefetch_related('batches').filter(is_active=True)
        now = timezone.now().date()

        total_items = stocks.count()
        low_stock_count = sum(1 for s in stocks if s.is_low_stock)
        out_of_stock = sum(1 for s in stocks if s.total_quantity == 0)

        # Stock valuation
        total_cost_value = 0
        total_retail_value = 0
        for s in stocks:
            qty = s.total_quantity
            total_cost_value += float(s.cost_price) * qty
            total_retail_value += float(s.selling_price) * qty

        # Expiring within 30/90 days
        expiring_30 = StockBatch.objects.filter(
            expiry_date__lte=now + timedelta(days=30),
            expiry_date__gt=now,
            quantity_remaining__gt=0,
        ).count()
        expiring_90 = StockBatch.objects.filter(
            expiry_date__lte=now + timedelta(days=90),
            expiry_date__gt=now,
            quantity_remaining__gt=0,
        ).count()
        expired = StockBatch.objects.filter(
            expiry_date__lt=now,
            quantity_remaining__gt=0,
        ).count()

        return Response({
            'total_items': total_items,
            'low_stock_count': low_stock_count,
            'out_of_stock': out_of_stock,
            'total_cost_value': round(total_cost_value, 2),
            'total_retail_value': round(total_retail_value, 2),
            'potential_profit': round(total_retail_value - total_cost_value, 2),
            'expiring_30_days': expiring_30,
            'expiring_90_days': expiring_90,
            'expired_batches': expired,
        })


# ─────────────────────────────────────────────────────────────────────────
#  Stock Take
# ─────────────────────────────────────────────────────────────────────────
class InventoryCountViewSet(viewsets.ModelViewSet):
    queryset = InventoryCount.objects.select_related('branch', 'category', 'created_by', 'completed_by').prefetch_related('lines__stock__unit').all()
    serializer_class = InventoryCountSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'branch', 'category']
    search_fields = ['reference', 'name', 'notes']
    ordering_fields = ['created_at', 'completed_at']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)

    @action(detail=True, methods=['post'], url_path='generate-sheet')
    def generate_sheet(self, request, pk=None):
        """Populate count lines with current expected qty from active stocks."""
        count = self.get_object()
        if count.status not in (InventoryCount.Status.DRAFT, InventoryCount.Status.IN_PROGRESS):
            return Response({'detail': 'Sheet can only be generated while draft / in progress.'}, status=400)

        qs = MedicationStock.objects.filter(is_active=True).prefetch_related('batches')
        if count.category_id:
            qs = qs.filter(category_id=count.category_id)

        # Wipe old lines and rebuild
        count.lines.all().delete()
        lines = []
        for s in qs:
            lines.append(InventoryCountLine(
                count=count, stock=s,
                expected_quantity=s.total_quantity,
            ))
        InventoryCountLine.objects.bulk_create(lines)
        count.status = InventoryCount.Status.IN_PROGRESS
        count.save(update_fields=['status'])
        return Response(self.get_serializer(count).data)

    @action(detail=True, methods=['post'], url_path='save-counts')
    def save_counts(self, request, pk=None):
        """Bulk-update counted quantities. Payload: {lines: [{id, counted_quantity, notes}]}"""
        count = self.get_object()
        if count.status == InventoryCount.Status.COMPLETED:
            return Response({'detail': 'Count is completed and cannot be edited.'}, status=400)
        lines = request.data.get('lines') or []
        by_id = {l.id: l for l in count.lines.all()}
        for payload in lines:
            line = by_id.get(payload.get('id'))
            if not line:
                continue
            cq = payload.get('counted_quantity')
            line.counted_quantity = int(cq) if cq is not None and cq != '' else None
            line.notes = payload.get('notes', line.notes) or ''
            line.save(update_fields=['counted_quantity', 'notes'])
        return Response(self.get_serializer(count).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Finalise the count: create StockAdjustments for any variance."""
        count = self.get_object()
        if count.status == InventoryCount.Status.COMPLETED:
            return Response({'detail': 'Already completed.'}, status=400)

        adjustments_created = 0
        for line in count.lines.exclude(counted_quantity__isnull=True):
            variance = line.variance
            if variance == 0:
                continue
            stock = line.stock
            if variance > 0:
                # Add to most recent batch (or create a synthetic one)
                batch = stock.batches.order_by('-expiry_date').first()
                if batch is None:
                    from datetime import date as _d, timedelta as _td
                    batch = StockBatch.objects.create(
                        stock=stock,
                        batch_number=f'CNT-{count.pk}-{stock.pk}',
                        quantity_received=variance,
                        quantity_remaining=0,
                        cost_price_per_unit=stock.cost_price or 0,
                        expiry_date=_d.today() + _td(days=365),
                    )
                StockAdjustment.objects.create(
                    stock=stock, batch=batch, quantity_change=variance,
                    reason=StockAdjustment.Reason.COUNT_CORRECTION,
                    notes=f'Stock take {count.reference}',
                    adjusted_by=request.user if request.user.is_authenticated else None,
                )
                batch.quantity_remaining += variance
                batch.save(update_fields=['quantity_remaining'])
                adjustments_created += 1
            else:
                remaining = -variance
                for batch in stock.batches.filter(quantity_remaining__gt=0).order_by('expiry_date'):
                    if remaining <= 0:
                        break
                    take = min(batch.quantity_remaining, remaining)
                    StockAdjustment.objects.create(
                        stock=stock, batch=batch, quantity_change=-take,
                        reason=StockAdjustment.Reason.COUNT_CORRECTION,
                        notes=f'Stock take {count.reference}',
                        adjusted_by=request.user if request.user.is_authenticated else None,
                    )
                    batch.quantity_remaining -= take
                    batch.save(update_fields=['quantity_remaining'])
                    remaining -= take
                adjustments_created += 1

        count.status = InventoryCount.Status.COMPLETED
        count.completed_at = timezone.now()
        count.completed_by = request.user if request.user.is_authenticated else None
        count.save(update_fields=['status', 'completed_at', 'completed_by'])
        return Response({**self.get_serializer(count).data, 'adjustments_created': adjustments_created})


# ─────────────────────────────────────────────────────────────────────────
#  Stock Transfers
# ─────────────────────────────────────────────────────────────────────────
class StockTransferViewSet(viewsets.ModelViewSet):
    queryset = StockTransfer.objects.select_related(
        'source_branch', 'dest_branch', 'requested_by', 'approved_by', 'received_by'
    ).prefetch_related('lines__stock__unit').all()
    serializer_class = StockTransferSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'source_branch', 'dest_branch']
    search_fields = ['reference', 'notes']
    ordering_fields = ['requested_at', 'shipped_at', 'received_at']

    def _set_status(self, transfer, new_status, user, field):
        transfer.status = new_status
        if user and user.is_authenticated:
            setattr(transfer, field, user)
        if new_status == StockTransfer.Status.IN_TRANSIT:
            transfer.shipped_at = timezone.now()
        if new_status == StockTransfer.Status.COMPLETED:
            transfer.received_at = timezone.now()
        transfer.save()

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Move from draft to requested."""
        transfer = self.get_object()
        if transfer.status != StockTransfer.Status.DRAFT:
            return Response({'detail': 'Only drafts can be submitted.'}, status=400)
        transfer.status = StockTransfer.Status.REQUESTED
        transfer.save(update_fields=['status'])
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve and ship: deduct from source-branch stock immediately."""
        transfer = self.get_object()
        if transfer.status not in (StockTransfer.Status.REQUESTED, StockTransfer.Status.DRAFT):
            return Response({'detail': 'Only requested transfers can be approved.'}, status=400)

        # FEFO deduction at source branch (or any batch if branch unset)
        for line in transfer.lines.all():
            remaining = line.quantity
            batches = line.stock.batches.filter(
                quantity_remaining__gt=0,
            ).order_by('expiry_date')
            # Prefer batches at source branch first
            src_batches = list(batches.filter(branch=transfer.source_branch))
            other = [b for b in batches if b.branch_id != transfer.source_branch_id]
            for batch in src_batches + other:
                if remaining <= 0:
                    break
                take = min(batch.quantity_remaining, remaining)
                batch.quantity_remaining -= take
                batch.save(update_fields=['quantity_remaining'])
                remaining -= take

        self._set_status(transfer, StockTransfer.Status.IN_TRANSIT, request.user, 'approved_by')
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Receive at destination: create a new batch per line.

        Optional payload: {lines: [{id, quantity_received}]} to record short receipts.
        """
        transfer = self.get_object()
        if transfer.status != StockTransfer.Status.IN_TRANSIT:
            return Response({'detail': 'Only in-transit transfers can be received.'}, status=400)

        received_map = {}
        for payload in (request.data.get('lines') or []):
            received_map[payload.get('id')] = payload.get('quantity_received')

        from datetime import date as _d, timedelta as _td
        for line in transfer.lines.all():
            qr = received_map.get(line.id)
            qty_received = int(qr) if qr not in (None, '') else line.quantity
            line.quantity_received = qty_received
            line.save(update_fields=['quantity_received'])
            if qty_received > 0:
                StockBatch.objects.create(
                    stock=line.stock,
                    batch_number=f'TRF-{transfer.reference}',
                    quantity_received=qty_received,
                    quantity_remaining=qty_received,
                    cost_price_per_unit=line.stock.cost_price or 0,
                    expiry_date=_d.today() + _td(days=365),
                    branch=transfer.dest_branch,
                )

        self._set_status(transfer, StockTransfer.Status.COMPLETED, request.user, 'received_by')
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        transfer = self.get_object()
        if transfer.status in (StockTransfer.Status.COMPLETED, StockTransfer.Status.CANCELLED):
            return Response({'detail': 'Cannot cancel this transfer.'}, status=400)
        transfer.status = StockTransfer.Status.CANCELLED
        transfer.save(update_fields=['status'])
        return Response(self.get_serializer(transfer).data)


# ─────────────────────────────────────────────────────────────────────────
#  Controlled Substance Register
# ─────────────────────────────────────────────────────────────────────────
class ControlledSubstanceLogViewSet(viewsets.ModelViewSet):
    queryset = ControlledSubstanceLog.objects.select_related('recorded_by').all()
    serializer_class = ControlledSubstanceLogSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['action', 'medication_name', 'schedule']
    search_fields = ['medication_name', 'patient_name', 'patient_id_number',
                     'prescriber_name', 'prescription_reference', 'batch_number']
    ordering_fields = ['created_at', 'medication_name']

    def get_queryset(self):
        qs = super().get_queryset()
        df = self.request.query_params.get('date_from')
        dt = self.request.query_params.get('date_to')
        if df:
            qs = qs.filter(created_at__date__gte=df)
        if dt:
            qs = qs.filter(created_at__date__lte=dt)
        return qs

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user if self.request.user.is_authenticated else None)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        from django.db.models import Sum
        qs = self.get_queryset()
        by_action = {}
        for row in qs.values('action').annotate(c=Sum('quantity')):
            by_action[row['action']] = float(row['c'] or 0)
        by_med = list(
            qs.values('medication_name')
              .annotate(total=Sum('quantity'))
              .order_by('-total')[:10]
        )
        return Response({
            'total_records': qs.count(),
            'by_action': by_action,
            'top_medications': by_med,
        })

